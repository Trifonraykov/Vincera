"""Spreadsheet scanner: reads ONLY headers (row 1). Never reads data rows."""

from __future__ import annotations

import csv
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

from pydantic import BaseModel

if TYPE_CHECKING:
    from vincera.core.llm import OpenRouterClient

logger = logging.getLogger(__name__)


class SpreadsheetInfo(BaseModel):
    """Metadata about a spreadsheet file."""

    file_path: str
    file_name: str
    sheet_names: list[str] = []
    headers: list[str] = []
    estimated_row_count: int = 0
    file_size: int = 0
    last_modified: str = ""


class SpreadsheetScanner:
    """Scans spreadsheet headers only. Never reads data rows."""

    def __init__(self, llm: "OpenRouterClient") -> None:
        self._llm = llm

    async def scan_headers(
        self,
        file_paths: list[Path],
        max_files: int = 50,
    ) -> list[SpreadsheetInfo]:
        """Scan headers from spreadsheet files. Most recent first."""
        # Sort by modification time (most recent first)
        sorted_paths: list[Path] = []
        for p in file_paths:
            try:
                sorted_paths.append(p)
            except (OSError, ValueError):
                continue
        sorted_paths.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)

        results: list[SpreadsheetInfo] = []
        for path in sorted_paths[:max_files]:
            try:
                info = self._scan_single(path)
                if info:
                    results.append(info)
            except Exception as exc:
                logger.warning("Failed to scan %s: %s", path, exc)
                continue

        return results

    def _scan_single(self, path: Path) -> SpreadsheetInfo | None:
        """Scan a single spreadsheet file for headers."""
        suffix = path.suffix.lower()
        stat = path.stat()
        base = {
            "file_path": str(path),
            "file_name": path.name,
            "file_size": stat.st_size,
            "last_modified": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        }

        if suffix in (".xlsx", ".xls"):
            return self._scan_xlsx(path, base)
        elif suffix in (".csv", ".tsv"):
            return self._scan_csv(path, base, delimiter="\t" if suffix == ".tsv" else ",")
        return None

    def _scan_xlsx(self, path: Path, base: dict) -> SpreadsheetInfo | None:
        """Scan xlsx file: row 1 only."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            headers: list[str] = []
            row_count = 0

            ws = wb.active
            if ws is not None:
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(c) if c is not None else "" for c in row]
                row_count = ws.max_row or 0
                if row_count > 0:
                    row_count -= 1  # Subtract header row

            wb.close()
            return SpreadsheetInfo(
                **base,
                sheet_names=sheet_names,
                headers=headers,
                estimated_row_count=row_count,
            )
        except ImportError:
            logger.info("openpyxl not installed, skipping xlsx scanning")
            return None
        except Exception as exc:
            logger.warning("Failed to read xlsx %s: %s", path, exc)
            return None

    def _scan_csv(self, path: Path, base: dict, delimiter: str = ",") -> SpreadsheetInfo | None:
        """Scan CSV/TSV file: first line only."""
        try:
            with open(path, "r", newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=delimiter)
                headers = next(reader, [])

                # Count remaining lines for row estimate
                row_count = sum(1 for _ in f)

            return SpreadsheetInfo(
                **base,
                sheet_names=[],
                headers=headers,
                estimated_row_count=row_count,
            )
        except Exception as exc:
            logger.warning("Failed to read csv %s: %s", path, exc)
            return None

    async def analyze_patterns(self, spreadsheets: list[SpreadsheetInfo]) -> dict:
        """Use LLM to analyze patterns across spreadsheet headers."""
        if not spreadsheets:
            return {"summary": "No spreadsheets found.", "patterns": []}

        header_summary = "\n".join(
            f"- {s.file_name}: {', '.join(s.headers[:15])}" for s in spreadsheets[:20]
        )

        try:
            result = await self._llm.think_structured(
                system_prompt=(
                    "Analyze these spreadsheet headers and identify business patterns. "
                    "What types of data are being tracked? What manual processes do they reveal?"
                ),
                user_message=f"Spreadsheet headers:\n{header_summary}",
                response_schema={
                    "type": "object",
                    "properties": {
                        "summary": {"type": "string"},
                        "patterns": {
                            "type": "array",
                            "items": {"type": "string"},
                        },
                        "related_files": {
                            "type": "array",
                            "items": {"type": "string"},
                        },
                    },
                    "required": ["summary", "patterns"],
                },
            )
            return result
        except Exception as exc:
            logger.warning("LLM pattern analysis failed: %s", exc)
            return {"summary": "Analysis unavailable.", "patterns": []}
